"""CLI E2E：跑範例資料夾、輸出 xlsx、檢查存在。"""
from pathlib import Path

from openpyxl import load_workbook

from recorder2xlsx.cli import run_cli


def test_cli_runs(sample_folder: Path, tmp_path: Path) -> None:
    out = tmp_path / "out.xlsx"
    code = run_cli([str(sample_folder), "-o", str(out), "--interval", "120"])
    assert code == 0
    assert out.is_file()
    wb = load_workbook(out)
    assert "資料" in wb.sheetnames
    assert "事件" in wb.sheetnames
    assert wb["資料"].cell(4, 1).value is not None
