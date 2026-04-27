"""xlsx_writer 測試。"""
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook

from recorder2xlsx.core.models import ResampledRow
from recorder2xlsx.core.xlsx_writer import write_xlsx
from recorder2xlsx.format.alarm_log import Event
from recorder2xlsx.format.tag_cfg import Channel


def _ch(name: str) -> Channel:
    return Channel(name=name, unit="°C", range_low=-270.0, range_high=1370.0)


def test_basic_xlsx(tmp_path: Path) -> None:
    out = tmp_path / "out.xlsx"
    rows = [
        ResampledRow(
            timestamp=datetime(2026, 4, 24, 15, 47, 14, 183_000),
            values=["22.8", "24.2", "斷線"],
        ),
        ResampledRow(
            timestamp=datetime(2026, 4, 24, 15, 49, 14, 183_000),
            values=["22.8", "24.4", "23.8"],
        ),
    ]
    channels = [_ch("AI1"), _ch("AI2"), _ch("AI4")]
    events = [
        Event(
            action="開機",
            source="",
            occurred_at=datetime(2026, 4, 27, 14, 1, 14),
            cleared_at=None,
            value="",
            acknowledged=False,
        )
    ]
    write_xlsx(
        out,
        rows=rows,
        channels=channels,
        events=events,
        selected_channel_names=["AI1", "AI2", "AI4"],
        include_events=True,
    )

    wb = load_workbook(out)
    assert wb.sheetnames == ["資料", "事件"]

    data = wb["資料"]
    assert data.cell(1, 1).value == "通道"
    assert data.cell(1, 2).value == "時間"
    assert data.cell(1, 3).value == "AI1"
    assert data.cell(1, 5).value == "AI4"
    assert data.cell(2, 3).value == "°C"
    assert data.cell(3, 3).value == "-270.0~1370.0"
    assert data.cell(4, 1).value == "2026-04-24"
    assert data.cell(4, 2).value == "15:47:14:183"
    assert data.cell(4, 3).value == 22.8
    assert data.cell(4, 5).value == "斷線"
    assert data.cell(5, 5).value == 23.8

    ev = wb["事件"]
    assert ev.cell(1, 2).value == "動作"
    assert ev.cell(2, 2).value == "開機"


def test_skip_events(tmp_path: Path) -> None:
    out = tmp_path / "out.xlsx"
    write_xlsx(
        out,
        rows=[],
        channels=[_ch("AI1")],
        events=[],
        selected_channel_names=["AI1"],
        include_events=False,
    )
    wb = load_workbook(out)
    assert wb.sheetnames == ["資料"]
