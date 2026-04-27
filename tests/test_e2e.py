"""E2E：對範例資料夾跑 CLI，輸出 xlsx 與 Windows 範例 CSV 逐欄比對。"""
import csv
from datetime import datetime
from pathlib import Path

import pytest
from openpyxl import load_workbook

from recorder2xlsx.cli import run_cli


def test_xlsx_matches_windows_csv(
    sample_folder: Path, expected_pen_csv: Path, tmp_path: Path
) -> None:
    out = tmp_path / "out.xlsx"
    code = run_cli([str(sample_folder), "-o", str(out), "--interval", "120"])
    assert code == 0

    wb = load_workbook(out)
    ws = wb["資料"]

    with expected_pen_csv.open("r", encoding="big5") as f:
        reader = csv.reader(f)
        next(reader)  # 通道
        next(reader)  # 量測方式
        next(reader)  # 工程單位
        expected_rows = [r for r in reader if r and r[0]]

    assert len(expected_rows) == 709

    for i, exp in enumerate(expected_rows):
        row_num = i + 4
        exp_date = datetime.strptime(exp[0], "%m/%d/%Y").strftime("%Y-%m-%d")
        assert ws.cell(row_num, 1).value == exp_date
        assert ws.cell(row_num, 2).value == exp[1]
        for col, exp_val in enumerate(exp[2:20], start=3):
            actual = ws.cell(row_num, col).value
            if exp_val == "錯誤":
                assert actual == "斷線", (
                    f"row {row_num} col {col}: 期望 '斷線'，實際 '{actual}'"
                )
            elif exp_val.strip() == "":
                assert actual in (None, "")
            else:
                assert actual == pytest.approx(float(exp_val), abs=0.05), (
                    f"row {row_num} col {col}: 期望 {exp_val}，實際 {actual}"
                )
