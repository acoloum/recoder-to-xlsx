"""寫 .xlsx：資料分頁（含 3 列表頭）+ 事件分頁。"""
from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from openpyxl.utils import get_column_letter

from ..format.alarm_log import Event
from ..format.tag_cfg import Channel
from .models import ResampledRow


def write_xlsx(
    path: Path,
    *,
    rows: list[ResampledRow],
    channels: list[Channel],
    events: list[Event],
    selected_channel_names: list[str],
    include_events: bool,
) -> None:
    wb = Workbook()
    data_ws = wb.active
    data_ws.title = "資料"

    # row 1：通道名稱
    data_ws.cell(1, 1, "通道")
    data_ws.cell(1, 2, "時間")
    name_to_ch = {c.name: c for c in channels}
    for col_idx, name in enumerate(selected_channel_names, start=3):
        data_ws.cell(1, col_idx, name)

    # row 2：工程單位
    for col_idx, name in enumerate(selected_channel_names, start=3):
        data_ws.cell(2, col_idx, name_to_ch[name].unit)

    # row 3：量測範圍
    for col_idx, name in enumerate(selected_channel_names, start=3):
        ch = name_to_ch[name]
        data_ws.cell(3, col_idx, f"{ch.range_low}~{ch.range_high}")

    # row 4+：資料
    for row_offset, row in enumerate(rows, start=4):
        ts = row.timestamp
        data_ws.cell(row_offset, 1, ts.strftime("%Y-%m-%d"))
        ms = ts.microsecond // 1000
        data_ws.cell(row_offset, 2, ts.strftime("%H:%M:%S:") + f"{ms:03d}")
        for col_idx, val in enumerate(row.values, start=3):
            try:
                data_ws.cell(row_offset, col_idx, float(val))
            except (TypeError, ValueError):
                data_ws.cell(row_offset, col_idx, val)

    data_ws.freeze_panes = "C4"
    for col in range(1, 2 + len(selected_channel_names) + 1):
        data_ws.column_dimensions[get_column_letter(col)].width = 14

    if include_events:
        ev_ws = wb.create_sheet("事件")
        ev_ws.append(["確認", "動作", "來源", "發生時間", "清除時間", "數值"])
        for e in events:
            ev_ws.append([
                "是" if e.acknowledged else "",
                e.action,
                e.source,
                e.occurred_at.strftime("%Y-%m-%d %H:%M:%S"),
                e.cleared_at.strftime("%Y-%m-%d %H:%M:%S") if e.cleared_at else "",
                e.value,
            ])
        for col in range(1, 7):
            ev_ws.column_dimensions[get_column_letter(col)].width = 18

    wb.save(path)
